from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def crear_excel_formato_especifico(df, archivo_salida):
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos Procesados"

    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    header_font = Font(bold=True, color="000000", size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))
    alignment_header = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alignment_data = Alignment(vertical='center')

    rows = dataframe_to_rows(df, index=False, header=True)

    for row_idx, row in enumerate(rows, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = alignment_header
                cell.border = border
            else:
                cell.border = border
                cell.alignment = alignment_data

    for column in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

    ws.freeze_panes = "A2"

    wb.save(archivo_salida)


def crear_hoja_resumen_simple(wb, df):
    ws_resumen = wb.create_sheet("Resumen")
    info = [
        ["RESUMEN DEL REPORTE", ""],
        ["", ""],
        ["Total de Registros", len(df)],
        ["Total de Columnas", len(df.columns)],
        ["Fecha de Procesamiento", datetime.now().strftime('%d/%m/%Y %H:%M')],
        ["", ""],
        ["Columnas en el reporte:", ""]
    ]
    for i, columna in enumerate(df.columns, 1):
        info.append([f"  {i}. {columna}", ""])

    for row_idx, (label, value) in enumerate(info, 1):
        ws_resumen[f'A{row_idx}'] = label
        ws_resumen[f'B{row_idx}'] = value

    ws_resumen.column_dimensions['A'].width = 25
    ws_resumen.column_dimensions['B'].width = 20
